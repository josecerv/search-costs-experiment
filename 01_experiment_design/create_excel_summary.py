#!/usr/bin/env python3
"""
Create Excel summary of bit.ly click rates for bar graphs.
Calculates Email 1, Email 2, and Overall click rates with means and SEs.
"""

import csv
import json
import requests
import time
from datetime import datetime, timedelta
import math
from collections import defaultdict
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    has_excel = True
except ImportError:
    has_excel = False
    print("openpyxl not installed, will output CSV instead")

# Configuration
API_KEY = "cf0f12b521bd873f4ac558e366cc946b456ef90d"
EMAIL1_START = "2024-06-24"
EMAIL1_END = "2024-07-09"
EMAIL2_START = "2024-07-09"
ANALYSIS_END = "2024-08-23"

def get_bitly_clicks_by_date(bitlink, api_key, start_date, end_date):
    """Fetch daily click data for a specific bit.ly link."""
    if "bit.ly/" in bitlink:
        short_code = bitlink.split("bit.ly/")[1]
        bitlink_id = f"bit.ly/{short_code}"
    else:
        return []
    
    url = f"https://api-ssl.bitly.com/v4/bitlinks/{bitlink_id}/clicks"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    days_diff = (end_date - start_date).days
    params = {
        "unit": "day",
        "units": days_diff,
        "unit_reference": end_date.strftime("%Y-%m-%dT%H:%M:%S+0000")
    }
    
    try:
        response = requests.get(url, headers=headers, params=params)
        if response.status_code == 200:
            data = response.json()
            return data.get("link_clicks", [])
        else:
            return []
    except Exception:
        return []

def aggregate_clicks_by_period(click_data, email1_end_date):
    """Aggregate clicks into email 1 and email 2 periods."""
    email1_clicks = 0
    email2_clicks = 0
    
    for entry in click_data:
        click_date = datetime.strptime(entry['date'][:10], "%Y-%m-%d")
        clicks = entry['clicks']
        
        if click_date < email1_end_date:
            email1_clicks += clicks
        else:
            email2_clicks += clicks
    
    return email1_clicks, email2_clicks

def calculate_stats(clicked_list):
    """Calculate mean and SE for a list of 0s and 1s."""
    n = len(clicked_list)
    if n == 0:
        return {'mean': 0, 'se': 0, 'mean_pct': 0, 'se_pct': 0, 'n': 0, 'clicked': 0}
    
    clicked = sum(clicked_list)
    mean = clicked / n
    # Standard error for proportion
    se = math.sqrt(mean * (1 - mean) / n) if n > 0 else 0
    
    return {
        'mean': mean,
        'se': se,
        'mean_pct': mean * 100,
        'se_pct': se * 100,
        'n': n,
        'clicked': clicked
    }

def main():
    print("Loading data...")
    
    # Load email campaign data
    emails = []
    with open("/mnt/c/Users/jcerv/Jose/search-costs/02_intervention_materials/email_campaigns/email-launch.csv", 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            emails.append(row)
    
    # Date ranges
    email1_start = datetime.strptime(EMAIL1_START, "%Y-%m-%d")
    email1_end = datetime.strptime(EMAIL1_END, "%Y-%m-%d")
    analysis_end = datetime.strptime(ANALYSIS_END, "%Y-%m-%d")
    
    # Get unique links
    unique_links = list(set(row['Link'] for row in emails))
    
    # Load cached click data if available
    cache_file = '/mnt/c/Users/jcerv/Jose/search-costs/01_experiment_design/bitly_cache.json'
    try:
        with open(cache_file, 'r') as f:
            link_clicks_by_period = json.load(f)
        print("Loaded click data from cache")
    except:
        print("Fetching click data from bit.ly API...")
        link_clicks_by_period = {}
        
        for i, link in enumerate(unique_links):
            if i % 10 == 0:
                print(f"Progress: {i}/{len(unique_links)} links processed")
            
            click_data = get_bitly_clicks_by_date(link, API_KEY, email1_start, analysis_end)
            email1_clicks, email2_clicks = aggregate_clicks_by_period(click_data, email1_end)
            
            link_clicks_by_period[link] = {
                'email1': email1_clicks,
                'email2': email2_clicks,
                'total': email1_clicks + email2_clicks
            }
            time.sleep(0.1)
        
        # Save cache
        with open(cache_file, 'w') as f:
            json.dump(link_clicks_by_period, f)
    
    # Map clicks to unique seminars
    print("Analyzing click data by seminar...")
    
    seminar_data = defaultdict(lambda: {
        'links': set(), 
        'condition': None, 
        'email1_clicked': 0, 
        'email2_clicked': 0,
        'overall_clicked': 0
    })
    
    # Process each email record
    for row in emails:
        link = row['Link']
        condition = row['condition']
        seminar_names = [s.strip() for s in row['seminar_names'].split(',')]
        
        for sem_name in seminar_names:
            dept = row['department']
            sem_key = f"{dept}|{sem_name}"
            
            seminar_data[sem_key]['links'].add(link)
            seminar_data[sem_key]['condition'] = condition
    
    # Calculate clicks for each seminar
    for sem_key, data in seminar_data.items():
        # Check if ANY of the seminar's links were clicked in each period
        for link in data['links']:
            clicks = link_clicks_by_period.get(link, {})
            if clicks.get('email1', 0) > 0:
                data['email1_clicked'] = 1
            if clicks.get('email2', 0) > 0:
                data['email2_clicked'] = 1
            if clicks.get('email1', 0) > 0 or clicks.get('email2', 0) > 0:
                data['overall_clicked'] = 1
    
    # Separate by condition
    control_email1 = []
    control_email2 = []
    control_overall = []
    treatment_email1 = []
    treatment_email2 = []
    treatment_overall = []
    
    for sem_key, data in seminar_data.items():
        if data['condition'] == 'control':
            control_email1.append(data['email1_clicked'])
            control_email2.append(data['email2_clicked'])
            control_overall.append(data['overall_clicked'])
        else:
            treatment_email1.append(data['email1_clicked'])
            treatment_email2.append(data['email2_clicked'])
            treatment_overall.append(data['overall_clicked'])
    
    # Calculate statistics
    results = {
        'Email 1': {
            'Control': calculate_stats(control_email1),
            'Treatment': calculate_stats(treatment_email1)
        },
        'Email 2': {
            'Control': calculate_stats(control_email2),
            'Treatment': calculate_stats(treatment_email2)
        },
        'Overall': {
            'Control': calculate_stats(control_overall),
            'Treatment': calculate_stats(treatment_overall)
        }
    }
    
    # Print results
    print("\n" + "="*60)
    print("SUMMARY FOR EXCEL BAR GRAPHS")
    print("="*60)
    
    for period in ['Email 1', 'Email 2', 'Overall']:
        print(f"\n{period}:")
        print(f"  Control: {results[period]['Control']['mean_pct']:.2f}% (SE: {results[period]['Control']['se_pct']:.2f}%)")
        print(f"  Treatment: {results[period]['Treatment']['mean_pct']:.2f}% (SE: {results[period]['Treatment']['se_pct']:.2f}%)")
    
    # Create Excel file if possible
    if has_excel:
        wb = Workbook()
        ws = wb.active
        ws.title = "Click Rate Summary"
        
        # Headers
        headers = ['Period', 'Condition', 'N', 'Clicked', 'Click Rate (%)', 'SE (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        row = 2
        for period in ['Email 1', 'Email 2', 'Overall']:
            for condition in ['Control', 'Treatment']:
                data = results[period][condition]
                ws.cell(row=row, column=1, value=period)
                ws.cell(row=row, column=2, value=condition)
                ws.cell(row=row, column=3, value=data['n'])
                ws.cell(row=row, column=4, value=data['clicked'])
                ws.cell(row=row, column=5, value=round(data['mean_pct'], 2))
                ws.cell(row=row, column=6, value=round(data['se_pct'], 2))
                row += 1
        
        # Adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Save
        excel_file = '/mnt/c/Users/jcerv/Jose/search-costs/01_experiment_design/bitly_click_summary.xlsx'
        wb.save(excel_file)
        print(f"\nExcel file saved: {excel_file}")
    
    # Also save as CSV
    csv_file = '/mnt/c/Users/jcerv/Jose/search-costs/01_experiment_design/bitly_click_summary.csv'
    with open(csv_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Period', 'Condition', 'N', 'Clicked', 'Click_Rate_Pct', 'SE_Pct'])
        for period in ['Email 1', 'Email 2', 'Overall']:
            for condition in ['Control', 'Treatment']:
                data = results[period][condition]
                writer.writerow([
                    period, condition, data['n'], data['clicked'],
                    round(data['mean_pct'], 2), round(data['se_pct'], 2)
                ])
    print(f"CSV file saved: {csv_file}")
    
    return results

if __name__ == "__main__":
    results = main()